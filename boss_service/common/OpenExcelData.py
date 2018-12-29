#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/12/13 0013 15:30
# @Site    : 
# @File    : OpenExcelData.py
# @Software: PyCharm
from openpyxl import load_workbook
import json
import os


def main(file):
    # contractPath = os.getcwd()
    # try:
    #     file.save(os.path.join(contractPath, file.filename))
    # except Exception as e:
    #     print e
    # wb = load_workbook(os.path.join(contractPath, file.filename), read_only=True)
    wb = load_workbook(file, read_only=True)
    sheet = wb.active
    allList = []
    data = ["deptUrl", "deptName", "desUrl", "evidenceUrl", "LeveLName", "levelCode", "deptClassifyName",
            "deptClassifyCode", "sectionCode", "deptAddress"]
    frist = True
    for row in sheet.rows:
        infoDict = {}
        if not frist:
            for index, cell in enumerate(row):
                infoDict[data[index]] = cell.value
            allList.append(infoDict)
        else:
            frist = False
    # print "\n".join(json.dumps(s) for s in allList)
    return allList
